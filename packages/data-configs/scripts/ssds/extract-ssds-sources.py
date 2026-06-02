#!/usr/bin/env python3
"""社会・人口統計体系 (SSDS) の cdCat01 → 原典調査 (資料源) マッピングを公式 Excel から抽出する。

背景:
  - SSDS は複数の原統計を加工編成した二次統計。e-Stat の JSON API (getMetaInfo /
    getStatsData) には「資料源 (原典統計)」フィールドが存在しないため、API では原典を取れない。
  - 総務省統計局が配布する 2 つの公式 Excel に原典が機械可読で入っている:
      * 基礎データ項目一覧   kiso_ken.xlsx    … 項目符号 → 「資料源」列 (原典統計名)
      * 社会生活統計指標一覧 shihyou_ken.xlsx … 指標コード → 「指標計算式」列 (基礎項目への分解式)
  - 指標コード (#A01202 等) は基礎項目の計算式なので、計算式に含まれる基礎項目コードを
    抽出し、それぞれの資料源を union することで原典を解決する (1 指標 ↔ 複数原典あり)。

出力:
  packages/data-configs/src/ssds/cdcat01-sources.generated.json … cdCat01 → {kind, sources[], formula?}
  packages/data-configs/src/ssds/source-names.generated.json     … 原典名のユニーク一覧 (正規化辞書の素材)

再生成 (総務省が年次更新したとき):
  python3 packages/data-configs/scripts/ssds/extract-ssds-sources.py
  ※ openpyxl 必須。Excel は stat.go.jp から自動ダウンロードする。

出典 (アクセス日 2026-06-02):
  https://www.stat.go.jp/data/ssds/2.html
  https://www.stat.go.jp/data/ssds/zuhyou/kiso_ken.xlsx
  https://www.stat.go.jp/data/ssds/zuhyou/shihyou_ken.xlsx
"""

from __future__ import annotations

import json
import re
import urllib.request
from pathlib import Path

KISO_URL = "https://www.stat.go.jp/data/ssds/zuhyou/kiso_ken.xlsx"
SHIHYOU_URL = "https://www.stat.go.jp/data/ssds/zuhyou/shihyou_ken.xlsx"

REPO_ROOT = Path(__file__).resolve().parents[4]
OUT_DIR = REPO_ROOT / "packages/data-configs/src/ssds"
CACHE_DIR = Path("/tmp/ssds")

# 基礎項目コード: 英大文字1 + 数字3〜7桁 (例 A1101 / B1103 / I821104)
BASE_RE = re.compile(r"[A-Z]\d{3,7}")


def download(url: str, dest: Path) -> Path:
    dest.parent.mkdir(parents=True, exist_ok=True)
    if not dest.exists():
        urllib.request.urlretrieve(url, dest)
    return dest


def parse_sources(cell: object) -> list[str]:
    """「国勢調査報告」、「人口推計」 のような資料源セルから原典名のリストを抽出。"""
    if not cell:
        return []
    return [m.strip() for m in re.findall(r"「([^」]+)」", str(cell))]


def main() -> None:
    import openpyxl  # 遅延 import (実行時のみ必須)

    kiso_path = download(KISO_URL, CACHE_DIR / "kiso_ken.xlsx")
    shihyou_path = download(SHIHYOU_URL, CACHE_DIR / "shihyou_ken.xlsx")

    # 1) 基礎項目: 項目符号 → [資料源]
    wb = openpyxl.load_workbook(kiso_path, read_only=True, data_only=True)
    ws = wb["基礎項目一覧（都道府県項目_基礎）"]
    kiso: dict[str, list[str]] = {}
    for row in list(ws.iter_rows(values_only=True))[4:]:
        code = str(row[1]).strip() if row[1] else ""
        if code and row[4]:
            kiso[code] = sorted(set(parse_sources(row[4])))

    # 2) 指標: 指標コード → 指標計算式
    wb2 = openpyxl.load_workbook(shihyou_path, read_only=True, data_only=True)
    ws2 = wb2["指標項目一覧（都道府県項目_指標）"]
    shihyou: dict[str, str] = {}
    for row in list(ws2.iter_rows(values_only=True))[4:]:
        code = str(row[1]).strip() if row[1] else ""
        if code.startswith("#"):
            shihyou[code] = str(row[4]).strip() if row[4] else ""

    def base_to_src(base: str) -> set[str]:
        if base in kiso:
            return set(kiso[base])
        # 末端へ丸める (接頭一致)
        for k in kiso:
            if base.startswith(k) or k.startswith(base):
                return set(kiso[k])
        return set()

    # 3) cdCat01 単位の安定マップ (基礎 + 指標)
    out: dict[str, dict] = {}
    for code, src in kiso.items():
        out[code] = {"kind": "base", "sources": src}
    for code, formula in shihyou.items():
        srcs: set[str] = set()
        for base in set(BASE_RE.findall(formula)):
            srcs |= base_to_src(base)
        out[code] = {"kind": "indicator", "formula": formula, "sources": sorted(srcs)}

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "cdcat01-sources.generated.json").write_text(
        json.dumps(out, ensure_ascii=False, indent=0), encoding="utf-8"
    )

    names: set[str] = set()
    for value in out.values():
        names |= set(value["sources"])
    (OUT_DIR / "source-names.generated.json").write_text(
        json.dumps(sorted(names), ensure_ascii=False, indent=1), encoding="utf-8"
    )

    print(f"cdCat01 entries: {len(out)} (base={len(kiso)} / indicator={len(shihyou)})")
    print(f"distinct source names: {len(names)}")


if __name__ == "__main__":
    main()

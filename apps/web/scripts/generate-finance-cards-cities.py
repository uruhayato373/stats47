"""市区町村決算カード (総務省) → per-prefecture city JSON + 類似団体平均 一括取込。
card-21..card-25.html (令和2..6 = 2020..2024) の都道府県別xlsxを取得し、各団体シートから抽出。
値列: 左ブロック=CO列, 右ブロック=CS列。歳入内訳=M列, 目的別歳出=BF列, 類型=「市町村類型」右。
出力:
  apps/web/public/finance-cards/cities/<prefCode>.json
     = { cityName: { type, years:{year:{12指標}}, flow:{revenue[],expenditure[],totals} } }
  apps/web/public/finance-cards/similar-averages.json
     = { 類型: { year: {12指標 平均} } }   ← 全国の同類型平均 = 類似団体平均
"""
import openpyxl, re, json, os, urllib.request, time
from openpyxl.utils import column_index_from_string as ci

YEAR_PAGE = {2020:"card-21",2021:"card-22",2022:"card-23",2023:"card-24",2024:"card-25"}
BASE = "https://www.soumu.go.jp"
TMP = "/tmp/citycards"
OUT = "/home/user/stats47/apps/web/public/finance-cards/cities"
OUT_SIM = "/home/user/stats47/apps/web/public/finance-cards/similar-averages.json"
LATEST = 2024
os.makedirs(TMP, exist_ok=True); os.makedirs(OUT, exist_ok=True)

PREFS = ["北海道","青森県","岩手県","宮城県","秋田県","山形県","福島県","茨城県","栃木県","群馬県","埼玉県","千葉県","東京都","神奈川県","新潟県","富山県","石川県","福井県","山梨県","長野県","岐阜県","静岡県","愛知県","三重県","滋賀県","京都府","大阪府","兵庫県","奈良県","和歌山県","鳥取県","島根県","岡山県","広島県","山口県","徳島県","香川県","愛媛県","高知県","福岡県","佐賀県","長崎県","熊本県","大分県","宮崎県","鹿児島県","沖縄県"]
PCODE = {p:f"{i+1:02d}" for i,p in enumerate(PREFS)}
UA = {"User-Agent":"Mozilla/5.0"}
def http(url): return urllib.request.urlopen(urllib.request.Request(url, headers=UA), timeout=90).read()
def norm(s): return re.sub(r"\s+","",str(s)) if s is not None else ""

CO = ci("CO"); CS = ci("CS"); M = ci("M"); BF = ci("BF")
LEFT = {"歳入総額":"revenue","歳出総額":"expenditure","実質収支":"realBalance"}
RIGHT = {"標準財政規模":"standardScale","財政力指数":"fiscalIndex","実質公債費比率":"debtServiceRatio","将来負担比率":"futureBurdenRatio","地方債現在高":"localDebt"}
# 歳入財源 (B列ラベル -> M列金額)
REV = {"地方税":"地方税","地方譲与税":"地方譲与税","地方交付税":"地方交付税","国庫支出金":"国庫支出金","都道府県支出金":"都道府県支出金","地方債":"地方債"}
# 目的別歳出 (AW列ラベル -> BF列金額)
EXP = ["議会費","総務費","民生費","衛生費","労働費","農林水産業費","商工費","土木費","消防費","教育費","災害復旧費","公債費","諸支出金"]
CORE = list(LEFT.values())+list(RIGHT.values())+["fundAdjust","fundRedemption","fundOther","currentBalanceRatio"]

def parse_sheet(ws):
    cells = {}; labels = []
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if v is None: continue
            cells[(c.row, c.column)] = v
            if isinstance(v, str):
                n = norm(v)
                if n: labels.append((c.row, c.column, n, v))
    def get(r, col): return cells.get((r, col))
    out = {}; fund_row = None; keiyo = None; ctype = None
    rev = {}; exp = {}
    for r, col, n, raw in labels:
        if n in LEFT and LEFT[n] not in out:
            v = get(r, CO)
            if isinstance(v,(int,float)): out[LEFT[n]] = v
        elif n in RIGHT and RIGHT[n] not in out:
            v = get(r, CS)
            if isinstance(v,(int,float)): out[RIGHT[n]] = v
        elif n == "積立金現在高" and fund_row is None:
            fund_row = r
        elif n == "経常収支比率" and keiyo is None:
            for dr in (0,1):
                for dc in range(1,9):
                    v = get(r+dr, col+dc)
                    if isinstance(v,(int,float)) and 40<=v<=150: keiyo = v; break
                if keiyo is not None: break
        elif n == "市町村類型" and ctype is None:
            for dc in range(1,16):
                v = get(r, col+dc)
                if isinstance(v,str) and v.strip(): ctype = v.strip(); break
        if n in REV and col <= 20 and REV[n] not in rev:
            v = get(r, M)
            if isinstance(v,(int,float)): rev[REV[n]] = v
        if n in EXP and 44 <= col <= 56 and n not in exp:
            v = get(r, BF)
            if isinstance(v,(int,float)): exp[n] = v
    if fund_row is not None:
        for key,dr in (("fundAdjust",0),("fundRedemption",1),("fundOther",2)):
            v = get(fund_row+dr, CS); out[key] = v if isinstance(v,(int,float)) else 0
    if keiyo is not None: out["currentBalanceRatio"] = keiyo
    return out, ctype, rev, exp

def build_flow(rev, exp, total_rev, total_exp):
    rev_nodes = [{"name":k,"value":v} for k,v in rev.items() if v>0]
    known = sum(n["value"] for n in rev_nodes)
    if total_rev and total_rev-known > 0: rev_nodes.append({"name":"その他財源","value":total_rev-known})
    exp_nodes = [{"name":k,"value":v} for k,v in exp.items() if v>0]
    keo = sum(n["value"] for n in exp_nodes)
    if total_exp and total_exp-keo > 0: exp_nodes.append({"name":"その他歳出","value":total_exp-keo})
    rev_nodes.sort(key=lambda n:-n["value"]); exp_nodes.sort(key=lambda n:-n["value"])
    return {"revenue":rev_nodes,"expenditure":exp_nodes,"totals":{"revenue":total_rev or 0,"expenditure":total_exp or 0}}

result = {}   # prefCode -> cityName -> {type, years:{}, flow}
sim_acc = {}  # type -> year -> {field: [values]}

def scrape_year(year, slug):
    html = http(f"{BASE}/iken/zaisei/{slug}.html").decode("shift_jis","ignore")
    cur = None; files = []
    prefs = "|".join(PREFS)
    for m in re.finditer(rf"({prefs})|href=\"(/main_content/\d+\.xlsx)\"", html):
        if m.group(1): cur = m.group(1)
        elif m.group(2) and cur: files.append((PCODE[cur], BASE+m.group(2)))
    print(f"[{year}] files: {len(files)}", flush=True)
    for pc, url in files:
        fn = f"{TMP}/{year}-{url.split('/')[-1]}"
        if not os.path.exists(fn):
            try: open(fn,"wb").write(http(url))
            except Exception: print("DL FAIL", url, flush=True); continue
        try: wb = openpyxl.load_workbook(fn, data_only=True, read_only=True)
        except Exception as e: print("OPEN FAIL", fn, e, flush=True); continue
        for sn in wb.sheetnames:
            if sn == "目次": continue
            name = re.sub(r"^\d+","", sn).strip()
            rec, ctype, rev, exp = parse_sheet(wb[sn])
            if not rec: continue
            city = result.setdefault(pc, {}).setdefault(name, {"type":ctype, "years":{}})
            if ctype and not city["type"]: city["type"] = ctype
            for k in CORE: rec.setdefault(k, 0)
            city["years"][str(year)] = rec
            if year == LATEST:
                city["flow"] = build_flow(rev, exp, rec.get("revenue"), rec.get("expenditure"))
            # 類似団体平均の集計
            t = ctype or city.get("type")
            if t:
                d = sim_acc.setdefault(t, {}).setdefault(year, {})
                for k in CORE: d.setdefault(k, []).append(rec.get(k,0))
        wb.close()

for y, slug in YEAR_PAGE.items():
    scrape_year(y, slug)

# 出力
tc = tr = 0
for pc, cities in result.items():
    json.dump(cities, open(f"{OUT}/{pc}.json","w"), ensure_ascii=False, separators=(",",":"))
    tc += len(cities); tr += sum(len(d["years"]) for d in cities.values())
sim = {}
for t, yrs in sim_acc.items():
    sim[t] = {}
    for y, fields in yrs.items():
        sim[t][str(y)] = {k: (sum(v)/len(v) if v else 0) for k,v in fields.items()}
json.dump(sim, open(OUT_SIM,"w"), ensure_ascii=False, separators=(",",":"))
print(f"DONE prefs={len(result)} cities={tc} records={tr} types={len(sim)}", flush=True)
s = result.get("01",{}).get("札幌市",{})
sy = s.get("years",{}).get("2024",{})
print("札幌市 type=",s.get("type")," 民生費(flow)=", next((n['value'] for n in s.get('flow',{}).get('expenditure',[]) if n['name']=='民生費'), None), " 経常",sy.get("currentBalanceRatio"), flush=True)
print("類似団体(政令指定都市)2024 経常平均=", round(sim.get("政令指定都市",{}).get("2024",{}).get("currentBalanceRatio",0),1), flush=True)

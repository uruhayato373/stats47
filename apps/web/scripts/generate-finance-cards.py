# 都道府県決算カード (総務省) → finance-cards.json 生成スクリプト。
# 完全DBレス: Reference データ (総務省公開Excelから再生成可能)。
# 取得元: https://www.soumu.go.jp/iken/zaisei/card.html の都道府県決算カード xlsx (年度別)。
# 使い方: 各年度 xlsx を /tmp/kessan-<year>.xlsx に置き `python3 generate-finance-cards.py`。
#   出力: apps/web/src/features/local-finance-dashboard/data/finance-cards.json
# 抽出: 各県シートの CQ列(当年度)。実質収支額/積立金内訳(財調/減債/その他)/地方債現在高/
#       標準財政規模/歳入歳出総額/財政力指数/経常収支比率/実質公債費比率/将来負担比率。
# 依存: pip install openpyxl

import openpyxl, json, re, sys
from openpyxl.utils import column_index_from_string as ci

YEARS = {2020:"kessan-2020.xlsx",2021:"kessan-2021.xlsx",2022:"kessan-2022.xlsx",
         2023:"kessan-2023.xlsx",2024:"kessan-2024.xlsx"}

# CQ列(当年度) にある: ラベル(正規化) -> 出力キー
CQ_FIELDS = {
  "歳入総額":"revenue", "歳出総額":"expenditure", "実質収支":"realBalance",
  "標準財政規模":"standardScale", "財政力指数":"fiscalIndex",
  "実質公債費比率":"debtServiceRatio", "将来負担比率":"futureBurdenRatio",
  "財政調整基金":"fundAdjust", "減債基金":"fundRedemption", "その他特定目的基金":"fundOther",
  "地方債現在高":"localDebt",
}
CQ_COL = ci("CQ")  # 95 当年度
LABEL_MIN, LABEL_MAX = ci("CC"), ci("CI")  # 81..87 ラベル帯

def norm(s):
    return re.sub(r"\s+","", str(s)) if s is not None else ""

def parse_sheet(ws):
    out = {}
    # CQ系: ラベル帯でラベル一致 -> 同行 CQ
    for row in ws.iter_rows():
        for c in row:
            if c.column < LABEL_MIN or c.column > LABEL_MAX: continue
            n = norm(c.value)
            if n in CQ_FIELDS and CQ_FIELDS[n] not in out:
                v = ws.cell(row=c.row, column=CQ_COL).value
                if isinstance(v,(int,float)):
                    out[CQ_FIELDS[n]] = v
    # 経常収支比率: ラベル完全一致 -> 同行で右隣の最初の数値(0..200)
    for row in ws.iter_rows():
        for c in row:
            if norm(c.value) == "経常収支比率":
                for cc in range(c.column+1, c.column+10):
                    v = ws.cell(row=c.row, column=cc).value
                    if isinstance(v,(int,float)) and 0<=v<=200:
                        out["currentBalanceRatio"] = v; break
                if "currentBalanceRatio" in out: break
        if "currentBalanceRatio" in out: break
    return out

def pref_code(sheet_name):
    m = re.match(r"(\d+)", sheet_name)
    return f"{int(m.group(1))-1:02d}" if m else None  # 2北海道->01, 48沖縄->47

result = {}  # pref2 -> {name, years:{year:{...}}}
for year, fn in YEARS.items():
    wb = openpyxl.load_workbook(fn, data_only=True, read_only=False)
    for sn in wb.sheetnames:
        if sn == "目次": continue
        code = pref_code(sn)
        if not code: continue
        name = re.sub(r"^\d+","", sn)
        rec = parse_sheet(wb[sn])
        result.setdefault(code, {"name":name, "years":{}})
        result[code]["years"][str(year)] = rec
    wb.close()

json.dump(result, open("/tmp/finance-cards.json","w"), ensure_ascii=False)
# 検証出力
hok = result["01"]
print("北海道 name:", hok["name"])
for y in ["2020","2022","2024"]:
    r = hok["years"].get(y,{})
    oku = lambda k: round(r[k]/100000) if k in r else None
    print(f" {y}: 歳入{oku('revenue')} 歳出{oku('expenditure')} 実質収支{oku('realBalance')} 標準財政規模{oku('standardScale')} 地方債{oku('localDebt')}億 | 財政力{r.get('fiscalIndex')} 経常{r.get('currentBalanceRatio')} 実質公債費{r.get('debtServiceRatio')} 将来負担{r.get('futureBurdenRatio')} | 基金 財調{oku('fundAdjust')}/減債{oku('fundRedemption')}/他{oku('fundOther')}")
# 欠損チェック
missing=0
allkeys=set(CQ_FIELDS.values())|{"currentBalanceRatio"}
for code,d in result.items():
    for y,rec in d["years"].items():
        miss=allkeys-set(rec.keys())
        if miss: missing+=1; print("MISSING",code,d["name"],y,miss)
print("総レコード:", sum(len(d['years']) for d in result.values()), " 欠損レコード:", missing, " 県数:", len(result))
